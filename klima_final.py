#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Klima Aylık Tüketim Raporu - PPD Parser
Folkart Blu Çeşme Yönetimi İçin
"""

import re
import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import pandas as pd

# Regex used throughout the parser to identify daire columns
_DAIRE_COL_REGEX = re.compile(r"^DAIRE\s+(\d+)([A-F])?$", re.IGNORECASE)

# Helper constants for identifying ortak areas
_ORTAK_KEYWORDS = ['LOBI', 'YONETIM', 'FITNESS', 'MUTFAK', 'P.O', 'BAYBAYAN', 'RES']

class PPDRawParser:
    """PPD dosyalarını okumak, işlemek ve raporlamak için yardımcı sınıf.

    Tüm girişler pandas DataFrame formatında yönetilir; bu sayede
    düşük birim sürelerde binlerce satır ayrıştırılabilir.
    """

    def __init__(self) -> None:
        # Türkçe ay isimleri (bazı rapor başlıklarında gerekebilir)
        self.months_tr: Dict[int, str] = {
            1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN",
            5: "MAYIS", 6: "HAZİRAN", 7: "TEMMUZ", 8: "AĞUSTOS",
            9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
        }
        self.numara_mapping: Dict[str, str] = {}  # YENİ -> ESKİ mapping
        self.daire_sirasi: List[int] = []  # Daire okuma sırası
        self.load_daire_sirasi()
    
    def load_daire_sirasi(self, path: Optional[Union[str, Path]] = None) -> None:
        """Daire sırasını `daire_sirasi.txt` dosyasından okur.

        Eğer `path` verilirse oradaki dosyayı kullanır; yoksa modül
        dizinine bakar. Hata durumunda liste boş kalır ve uyarı basılır.
        """
        try:
            sira_file = Path(path) if path else Path(__file__).parent / "daire_sirasi.txt"
            if sira_file.exists():
                with open(sira_file, 'r', encoding='utf-8') as f:
                    self.daire_sirasi = [int(line.strip()) for line in f if line.strip()]
                print(f"Daire sırası yüklendi ({len(self.daire_sirasi)} daire)")
            else:
                print("⚠ daire_sirasi.txt dosyası bulunamadı (varsayılan sırası kullanılacak)")
        except Exception as e:
            print(f"Daire sırası yüklenemedi: {e}")
    
    def load_numara_mapping(self, ekim_file: Union[str, Path]) -> bool:
        """Ekim formatındaki CSV'den eski-yeni numara haritalamasını alır.

        Dosyada 10. satırdan sonraki veriler içerir. Hata olursa `False`
        döndürür ve `numara_mapping` aynı kalır.
        """
        try:
            with open(ekim_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=';')
                for row in list(reader)[9:]:
                    if len(row) < 2 or not row[0].strip() or not row[1].strip():
                        continue
                    self.numara_mapping[row[1].strip()] = row[0].strip()
            print(f"{len(self.numara_mapping)} numaralama eşleşmesi yüklendi")
            return True
        except Exception as e:
            print(f"Numara mapping yüklenemedi: {e}")
            return False
    
    def _is_daire_column(self, col_name: str) -> bool:
        """Verilen sütun adının daire/alan verisi içerip içermediğine bakar."""
        col = col_name.strip().upper()
        if _DAIRE_COL_REGEX.match(col):
            # yalnızca 1‑80 arası numaralardan oluşan gerçek daireler
            num = int(_DAIRE_COL_REGEX.match(col).group(1))
            return 1 <= num <= 80
        # sabit ortak isimler
        return any(keyword in col for keyword in _ORTAK_KEYWORDS)

    def _normalize_daire_name(self, col_name: str) -> str:
        """Hüc dergisindeki gibi orijinal sütun adını döner (boş bırakma yok)."""
        return col_name.strip()

    def parse_ppd_file(self, file_path: Union[str, Path]) -> pd.DataFrame:
        """PPD CSV dosyasını DataFrame'e dönüştürür ve toplamları hesaplar.

        * Pandas kullanarak tüm sayı dönüşümlerini vektörize eder,
          böylece düşük uçlu CPU'larda bile hızlı çalışır.
        * `DAIRE` önekli sütunları bulur ve sonra saatlik değerleri toplar.
        """
        print(f"PPD dosyası işleniyor: {file_path}")
        # pandas hızlı okuma
        raw = pd.read_csv(file_path, sep=';', header=6, encoding='utf-8-sig', low_memory=False)

        # sadece daire/ortak sütunları seç
        daire_cols = [c for c in raw.columns if self._is_daire_column(c)]
        if not daire_cols:
            raise ValueError("PPD dosyasında daire sütunu bulunamadı")
        print(f"{len(daire_cols)} sütun seçildi")

        # tüm değerleri sayıya çevir, eksikler 0 olsun
        df_vals = raw[daire_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
        totals: pd.Series = df_vals.sum(axis=0)

        # sonuç tablosunu oluştur
        records: List[Dict[str, Any]] = []
        for col, tot in totals.items():
            name = self._normalize_daire_name(col)
            num = self.extract_daire_number(name)
            typ = self.get_daire_type(name)
            records.append({
                'DAİRE_ADI': name,
                'DAİRE_NO': num,
                'TİP': typ,
                'AYLIK_TUKETIM_WH': tot,
                'AYLIK_TUKETIM_KWH': tot / 1000,
            })

        df = pd.DataFrame(records)

        # suit dairelerini numaraya göre grupla (ortaklar zaten ayrı)
        df_ortak = df[df['TİP'] == 'ORTAK']
        df_suit = df[df['TİP'] == 'SÜİT']
        if not df_suit.empty:
            grouped = df_suit.groupby('DAİRE_NO', as_index=False)[
                ['AYLIK_TUKETIM_WH', 'AYLIK_TUKETIM_KWH']
            ].sum()
            grouped['DAİRE_ADI'] = 'DAIRE ' + grouped['DAİRE_NO'].astype(str)
            grouped['TİP'] = 'SÜİT'
            df = pd.concat([grouped, df_ortak], ignore_index=True)
        else:
            df = df_ortak.copy()

        if self.numara_mapping:
            df['ESKİ_NUMARA'] = df['DAİRE_NO'].astype(str).map(self.numara_mapping).fillna('')
        else:
            df['ESKİ_NUMARA'] = ''

        print(f"{len(df)} kayıt hazır")
        return df
    
    def extract_daire_number(self, daire_name):
        """Daire adından numarası çıkar"""
        daire_name = daire_name.strip().upper()
        
        if 'LOBI' in daire_name:
            return 'LOBI'
        if 'YONETIM' in daire_name:
            return 'YONETIM'
        if 'FITNESS' in daire_name:
            return 'FITNESS'
        if 'MUTFAK' in daire_name or 'P.O' in daire_name or 'BAYBAYAN' in daire_name:
            return 'ORTAK'
        
        match = re.search(r'(\d+)', daire_name)
        if match:
            return int(match.group(1))
        
        return daire_name
    
    def get_daire_type(self, daire_name):
        """Daire tipi belirle"""
        daire_name = daire_name.strip().upper()
        
        if any(x in daire_name for x in ['LOBI', 'YONETIM', 'FITNESS', 'MUTFAK', 'P.O', 'BAYBAYAN', 'RES']):
            return 'ORTAK'
        
        return 'SÜİT'
    
    def create_summary(self, df):
        """İstatistikler oluştur"""
        summary = {
            'Toplam Alan': len(df),
            'Genel Aylık Toplam (kWh)': df['AYLIK_TUKETIM_KWH'].sum(),
            'Ortalama (kWh)': df['AYLIK_TUKETIM_KWH'].mean(),
            'En Yüksek (kWh)': df['AYLIK_TUKETIM_KWH'].max(),
            'En Düşük (kWh)': df['AYLIK_TUKETIM_KWH'].min(),
        }
        
        for dtype in df['TİP'].unique():
            subset = df[df['TİP'] == dtype]
            summary[f'{dtype} - Toplam (kWh)'] = subset['AYLIK_TUKETIM_KWH'].sum()
            summary[f'{dtype} - Sayı'] = len(subset)
        
        return summary
    
    def export_results(self, df, summary, month_year, output_dir=None):
        """CSV ve Excel'e kaydet

        output_dir: Kullanıcı tarafından seçilen klasör (varsayılan olarak çalışma dizini)
        """
        # Daire sırasını uygula
        if len(self.daire_sirasi) > 0:
            # Daire sırasına göre sort et (sadece integer daireler)
            df_sorted = pd.DataFrame()
            for daire_no in self.daire_sirasi:
                daire_match = df[df['DAİRE_NO'] == daire_no]
                if len(daire_match) > 0:
                    df_sorted = pd.concat([df_sorted, daire_match], ignore_index=True)
            
            # Kalanları (sırada olmayan - ORTAK alanlar) sonuna ekle
            used_daires = set(self.daire_sirasi)
            df_remaining = df[~df['DAİRE_NO'].isin(used_daires)]
            # ORTAK alanları isme göre sırala
            if len(df_remaining) > 0:
                df_remaining = df_remaining.sort_values('DAİRE_ADI').reset_index(drop=True)
            df = pd.concat([df_sorted, df_remaining], ignore_index=True)
        
        # Dosya adı - sabit olarak ısıtma_sogutma (Türkçe karakterler yerine ascii)
        csv_name = "ısıtma_sogutma.csv"
        xlsx_name = "ısıtma_sogutma.xlsx"
        
        # Eğer bir çıkış dizini belirtilmişse ona göre yolu oluştur
        if output_dir:
            csv_file = str(Path(output_dir) / csv_name)
            xlsx_file = str(Path(output_dir) / xlsx_name)
        else:
            csv_file = csv_name
            xlsx_file = xlsx_name
        
        # CSV
        print(f"\n💾 CSV kaydediliyor: {csv_file}")
        with open(csv_file, 'w', encoding='utf-8-sig') as f:
            f.write("FOLKART BLU ÇEŞME YÖNETİMİ\n")
            f.write("ISITMA/SOĞUTMA RAPORU\n\n")
        
        # Sadece DAİRE_ADI, TİP, TÜKETİM WH/KWH sütunlarını dahil et
        # (DAİRE_NO ve ESKİ_NUMARA kullanıcı tarafından istenmiyor)
        df_export = df[['DAİRE_ADI', 'TİP', 'AYLIK_TUKETIM_WH', 'AYLIK_TUKETIM_KWH']]
        
        df_export.to_csv(csv_file, mode='a', index=False, encoding='utf-8-sig')
        
        with open(csv_file, 'a', encoding='utf-8-sig') as f:
            f.write("\n\nÖZET İSTATİSTİKLERİ\n")
            for key, value in summary.items():
                if isinstance(value, float):
                    f.write(f"{key};{value:.2f}\n")
                else:
                    f.write(f"{key};{value}\n")
        
        # Excel
        print(f"💾 Excel kaydediliyor: {xlsx_file}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tüketim"
        
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        # Başlık satırlarındaki sütun sayısını ayarla
        col_count = 6 if 'ESKİ_NUMARA' in df.columns else 5
        ws.merge_cells(f'A{row}:{"ABCDEF"[col_count-1]}{row}')
        cell = ws[f'A{row}']
        cell.value = "FOLKART BLU ÇEŞME YÖNETİMİ"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 25
        row += 1
        
        ws.merge_cells(f'A{row}:{"ABCDEF"[col_count-1]}{row}')
        cell = ws[f'A{row}']
        cell.value = "ISITMA/SOĞUTMA RAPORU"
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # Başlık satırı (sadece ad, tip ve tüketim)
        headers = ['DAİRE ADI', 'TİP', 'TÜKETİM (Wh)', 'TÜKETİM (kWh)']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        row += 1
        
        # Veri satırları
        for _, data_row in df.iterrows():
            col = 1
            ws.cell(row=row, column=col).value = data_row['DAİRE_ADI']
            ws.cell(row=row, column=col).border = border
            col += 1
            
            ws.cell(row=row, column=col).value = data_row['TİP']
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            col += 1
            
            ws.cell(row=row, column=col).value = data_row['AYLIK_TUKETIM_WH']
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).number_format = '0'
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            col += 1
            
            ws.cell(row=row, column=col).value = data_row['AYLIK_TUKETIM_KWH']
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).number_format = '0.00'
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            
            row += 1
        
        # Özet
        row += 2
        ws.merge_cells(f'A{row}:{"ABCDEF"[col_count-1]}{row}')
        cell = ws[f'A{row}']
        cell.value = "ÖZET İSTATİSTİKLERİ"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        row += 1
        
        for key, value in summary.items():
            ws.cell(row=row, column=1).value = key
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).border = border
            
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).border = border
            if isinstance(value, float):
                ws.cell(row=row, column=2).number_format = '0.00'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            
            row += 1
        
        # Kolon genişlikleri (sadece 4 kolon)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        wb.save(xlsx_file)
        
        print(f"\n✅ TAMAMLANDI!")
        print(f"   • {csv_file}")
        print(f"   • {xlsx_file}")
        
        return csv_file, xlsx_file
    
    def load_subat_sayac_data(self, excel_file):
        """Şubat sayaç okumaları Excel dosyasından veri yükle ve formata dönüştür"""
        print(f"\n📊 Şubat Sayaç Okumaları yükleniyor: {excel_file}")
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Daire verilerini oku (Satır 10'dan başlıyor)
            sayac_data = {}
            for row_idx in range(10, 100):
                eski_no = ws.cell(row_idx, 2).value
                yeni_no = ws.cell(row_idx, 3).value
                durum = ws.cell(row_idx, 4).value
                tuketim = ws.cell(row_idx, 7).value
                
                # Eğer tüm veriler boş ise dur
                if eski_no is None and yeni_no is None:
                    break
                
                # Yeni numaraya göre depolamak daha iyi
                if yeni_no is not None:
                    try:
                        yeni_no = int(yeni_no) if isinstance(yeni_no, str) else yeni_no
                        tuketim = float(tuketim) if tuketim else 0
                        sayac_data[yeni_no] = {
                            'ESKİ_NO': eski_no,
                            'YENİ_NO': yeni_no,
                            'DURUM': durum,
                            'TUKETIM': tuketim
                        }
                    except:
                        pass
            
            print(f"✓ {len(sayac_data)} daire verisi yüklendi")
            return sayac_data
        
        except Exception as e:
            print(f"⚠ Şubat verileri yüklenemedi: {e}")
            return {}
    
    def export_sayac_format(self, df, sayac_data, month_year):
        """Sayaç formatında Excel raporu oluştur"""
        print(f"\n💾 Sayaç Formatı Excel kaydediliyor...")
        
        # "/" karakterini "_" ile değiştir (Windows uyumluluğu)
        safe_filename = month_year.replace(' / ', '_')
        xlsx_file = f"Klima_{safe_filename}_SAYAÇ_OKUMALARI.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sayaç Okumaları"
        
        # Stiller
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        subtitle_font = Font(color="FFFFFF", bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "FOLKART BLU ÇEŞME YÖNETİMİ"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1
        
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = month_year
        cell.font = subtitle_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "ISITMA/SOĞUTMA SAYAÇ TÜKETİMLERİ"
        cell.font = subtitle_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # Başlık satırı
        headers = ['ESKİ NUMARASI', 'YENİ NUMARASI', 'DURUM', 'ISITMA/SOĞUTMA', 'İLK OKUMA', 'SON OKUMA', 'TÜKETİM']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Veri satırları - daire sırasına göre yerleştir
        # Sırası olan daireleri önce göster
        daire_order = self.daire_sirasi if len(self.daire_sirasi) > 0 else sorted(sayac_data.keys())
        
        for daire_no in daire_order:
            if daire_no not in sayac_data:
                continue
                
            sayac = sayac_data[daire_no]
            
            # PPD'den gelen tüketimi bul
            tuketim_kw = 0
            if len(df) > 0:
                # DAİRE_NO ile eşleştir
                daire_match = df[df['DAİRE_NO'] == daire_no]
                if len(daire_match) > 0:
                    tuketim_kw = daire_match.iloc[0]['AYLIK_TUKETIM_KWH']
            
            # Sayaç formatında kullan, eğer yoksa PPD verisi kullan
            tuketim_val = sayac.get('TUKETIM', tuketim_kw)
            
            ws.cell(row=row, column=1).value = sayac['ESKİ_NO']
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row, column=2).value = sayac['YENİ_NO']
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row, column=3).value = sayac['DURUM']
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row, column=4).value = ""  # ISITMA/SOĞUTMA etiketi boş
            ws.cell(row=row, column=4).border = border
            
            ws.cell(row=row, column=5).value = ""  # İLK OKUMA
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
            
            ws.cell(row=row, column=6).value = ""  # SON OKUMA
            ws.cell(row=row, column=6).border = border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            
            ws.cell(row=row, column=7).value = tuketim_val if tuketim_val else ""
            ws.cell(row=row, column=7).border = border
            ws.cell(row=row, column=7).number_format = '0.00'
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
            
            row += 1
        
        # Genişlikleri ayarla
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        wb.save(xlsx_file)
        print(f"✓ Sayaç formatı: {xlsx_file}")
        
        return xlsx_file
